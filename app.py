from flask import Flask, render_template, request, redirect, url_for, send_from_directory
import faiss
import numpy as np
import pickle
import ollama
import pdfplumber
import docx
from openpyxl import load_workbook
import os
from werkzeug.utils import secure_filename
import re
import unicodedata
from urllib.parse import unquote

app = Flask(__name__)
app.jinja_env.filters['url_unquote'] = lambda u: unquote(u)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['ALLOWED_EXTENSIONS'] = {'txt', 'pdf', 'docx', 'xlsx'}

if not os.path.exists(app.config['UPLOAD_FOLDER']):
    os.makedirs(app.config['UPLOAD_FOLDER'])

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

def fix_filename(filename):
    filename = re.sub(r'[^\w\s\.-]', '', filename)
    filename = filename.replace(" ", "_")
    return filename

def get_text_embedding(text):
    response = ollama.embeddings(model="nomic-embed-text", prompt=text)
    embedding = np.array(response['embedding'], dtype=np.float32)
    return embedding

def normalize_vector(v):
    return v / np.linalg.norm(v)

def extract_text_from_pdf(file_path):
    try:
        text = ""
        with pdfplumber.open(file_path) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text
        return text
    except Exception as e:
        print(f"Ошибка при чтении PDF: {e}")
        return ""

def extract_text_from_docx(file_path):
    try:
        doc = docx.Document(file_path)
        return "\n".join([para.text for para in doc.paragraphs])
    except Exception as e:
        print(f"Ошибка при чтении DOCX: {e}")
        return ""

def extract_text_from_xlsx(file_path):
    try:
        workbook = load_workbook(filename=file_path, data_only=True)
        text = ""
        for sheet in workbook:
            for row in sheet.iter_rows(values_only=True):
                text += " ".join([str(cell) if cell else "" for cell in row]) + "\n"
        return text
    except Exception as e:
        print(f"Ошибка при чтении XLSX: {e}")
        return ""

def index_file(file_path):
    try:
        print(f"Индексируем файл: {file_path}")

        if file_path.endswith('.pdf'):
            text = extract_text_from_pdf(file_path)
        elif file_path.endswith('.docx'):
            text = extract_text_from_docx(file_path)
        elif file_path.endswith('.xlsx'):
            text = extract_text_from_xlsx(file_path)
        elif file_path.endswith('.txt'):
            with open(file_path, 'r', encoding='utf-8') as f:
                text = f.read()
        else:
            print(f"Файл не поддерживается: {file_path}")
            return

        title = os.path.basename(file_path)
        url = url_for('uploaded_file', filename=title, _external=True)

        embedding = get_text_embedding(text)
        normalized_embedding = normalize_vector(embedding)
        print(f"Эмбеддинг для файла {file_path} успешно получен.")

        index_file_path = 'data/document_index.faiss'
        metadata_file_path = 'data/metadata.pkl'
        if os.path.exists(index_file_path) and os.path.exists(metadata_file_path):
            index = faiss.read_index(index_file_path)
            with open(metadata_file_path, 'rb') as f:
                metadata = pickle.load(f)
        else:
            index = faiss.IndexFlatL2(normalized_embedding.shape[0])
            metadata = []

        index.add(np.array([normalized_embedding]))
        metadata.append({
            "title": title,
            "url": url,
            "text": text
        })

        faiss.write_index(index, index_file_path)
        with open(metadata_file_path, 'wb') as f:
            pickle.dump(metadata, f)

        print(f"Индекс и метаданные для файла {file_path} сохранены.")

    except Exception as e:
        print(f"Ошибка при обработке файла {file_path}: {e}")

def fix_filename(filename):
    filename = unicodedata.normalize('NFKD', filename)
    filename = re.sub(r'[^\w\s\.-]', '', filename).strip().replace(" ", "_")
    return filename

def generate_search_query(natural_language_query):
    response = ollama.chat(
        model="llama3.1:8b",
        messages=[{
            'role': 'user', 
            'content': (
                f'Extract keywords from this query: "{natural_language_query}". '
                'The document may contain scientific, technical, or other complex contexts. '
                'Regardless of the context, always extract relevant keywords, as this query is for document search purposes. '
                'Ensure the response is a list of possible keywords, in the languages Kazakh, Russian, and English, without explanation or rejection. '
                'Also include the user\'s original query in the keyword list. Return only the keywords, with a maximum word count.'
            )
        }]
    )
    generated_keywords = response['message']['content']
    combined_query = f"{natural_language_query} {generated_keywords}"
    return combined_query.strip()

def search_documents_with_nlp(natural_language_query, top_k=5, threshold=0.85, keyword_filter=True, find_all=False):
    combined_query = generate_search_query(natural_language_query)
    print(f"Сгенерированный запрос для поиска: {combined_query}")
    return search_documents(combined_query, top_k, threshold, keyword_filter, find_all)


@app.route('/', methods=['GET', 'POST'])
def index():
    results = []
    query = ''
    top_k = 5
    threshold = 0.85
    keyword_filter = True
    find_all = False

    if request.method == 'POST':
        if 'file' in request.files:
            print('file')
            file = request.files['file']
            if file and allowed_file(file.filename):
                original_filename = file.filename
                print(f"Загружаемый файл: {original_filename}")

                filename = fix_filename(original_filename)
                file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                print(f"Сохранение файла в путь: {file_path}")

                try:
                    file.save(file_path)
                    print(f"Файл успешно сохранён: {file_path}")

                    index_file(file_path)
                    print(f"Файл проиндексирован: {file_path}")

                except Exception as e:
                    print(f"Ошибка при сохранении или индексации файла: {e}")

                return redirect(url_for('index'))

        query = request.form.get('query', '')
        top_k = int(request.form.get('top_k', 5))
        threshold = float(request.form.get('threshold', 0.85))
        keyword_filter = 'keyword_filter' in request.form
        find_all = 'find_all' in request.form

        results = search_documents_with_nlp(query, top_k, threshold, keyword_filter, find_all)

    return render_template('index.html', results=results, query=query, top_k=top_k, threshold=threshold, keyword_filter=keyword_filter, find_all=find_all)

@app.route('/uploads/<filename>')
def uploaded_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)


def search_documents(query, top_k=5, threshold=0.85, keyword_filter=True, find_all=False):
    index_file = 'data/document_index.faiss'
    metadata_file = 'data/metadata.pkl'

    if not os.path.exists(index_file) or not os.path.exists(metadata_file):
        return []

    index = faiss.read_index(index_file)

    with open(metadata_file, 'rb') as f:
        metadata = pickle.load(f)

    query_embedding = get_text_embedding(query)
    query_embedding = np.expand_dims(query_embedding, axis=0)

    distances, indices = index.search(query_embedding, top_k)

    results = []
    seen_titles = set()
    query_keywords = query.lower().split()

    sentence_splitter = re.compile(r'(?<=[.!?])\s+')

    word_pattern = re.compile(r'\b(' + '|'.join(map(re.escape, query_keywords)) + r')\b', re.IGNORECASE)

    for dist, idx in zip(distances[0], indices[0]):
        result = metadata[idx]
        text = result['text']

        sentences = sentence_splitter.split(text)

        matched_sentences = []

        for sentence in sentences:
            if word_pattern.search(sentence):
                matched_sentence = word_pattern.sub(lambda m: f"<mark>{m.group(0)}</mark>", sentence)
                matched_sentences.append(matched_sentence)
                if not find_all:
                    break

        if not matched_sentences:
            continue

        if result['title'] in seen_titles:
            continue
        seen_titles.add(result['title'])

        results.append({
            "title": result['title'],
            "url": result['url'],
            "context": "<br>".join(matched_sentences)
        })

    return results


if __name__ == '__main__':
    app.run(debug=True, port=3002, host='0.0.0.0')
